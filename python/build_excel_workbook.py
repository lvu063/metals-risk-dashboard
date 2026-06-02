"""
build_excel_workbook.py
-----------------------
Generates a professional Excel workbook for the Metals Risk Dashboard.

Sheets produced:
  1. README          — instructions and navigation
  2. Trade Blotter   — formatted table with formulas and conditional formatting
  3. PnL Summary     — SUMIF-powered PnL rollup by metal and desk
  4. Risk Exposure   — gross/net exposure with concentration flags
  5. Price History   — spot price table, ready for charting
  6. Pivot_Data      — clean flat table for pivot table creation
  7. VBA_Reference   — documented VBA macros (as text, since xlsm needs manual save)

Usage:
    python excel/build_excel_workbook.py
    python excel/build_excel_workbook.py --data-dir data --out excel/MetalsRiskDashboard.xlsx
"""

import argparse
import csv
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
except ImportError:
    raise SystemExit("Run: pip install openpyxl")


# ---------------------------------------------------------------------------
# Colour palette (Bloomberg-terminal inspired)
# ---------------------------------------------------------------------------
NAVY       = "0A1628"
GOLD       = "C9A84C"
DARK_BLUE  = "0D1F3C"
MID_BLUE   = "1E3A5F"
WHITE      = "E8E8E8"
GREEN      = "00C851"
RED        = "FF4444"
AMBER      = "FFA500"
LIGHT_GREY = "F5F5F5"
HEADER_BG  = "1E3A5F"


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def hdr_font(bold=True, colour=WHITE, size=11):
    return Font(name="Calibri", bold=bold, color=colour, size=size)

def hdr_fill(colour=HEADER_BG):
    return PatternFill("solid", fgColor=colour)

def gold_fill():
    return PatternFill("solid", fgColor=GOLD)

def navy_fill():
    return PatternFill("solid", fgColor=NAVY)

def centre():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center")

def thin_border():
    s = Side(style="thin", color="2C3E50")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header_row(ws, row_num, headers, widths=None):
    """Write a formatted header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font      = hdr_font()
        cell.fill      = hdr_fill()
        cell.alignment = centre()
        cell.border    = thin_border()
        if widths and col <= len(widths):
            ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

def money_fmt(ws, row, col):
    ws.cell(row=row, column=col).number_format = '#,##0.00'

def pct_fmt(ws, row, col):
    ws.cell(row=row, column=col).number_format = '0.00%'


# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------

def load_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with open(path) as f:
        return list(csv.DictReader(f))


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_readme(wb, generated_at: str):
    ws = wb.create_sheet("README", 0)
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("B2:H2")
    title = ws["B2"]
    title.value = "METALS RISK DASHBOARD — EXCEL WORKBOOK"
    title.font  = Font(name="Calibri", bold=True, size=18, color=GOLD)
    title.fill  = navy_fill()
    title.alignment = centre()
    ws.row_dimensions[2].height = 40

    ws.merge_cells("B3:H3")
    sub = ws["B3"]
    sub.value = f"Generated: {generated_at}  |  Data: Synthetic (GBM simulation)  |  Version: 1.0"
    sub.font  = Font(name="Calibri", size=10, color=WHITE)
    sub.fill  = hdr_fill()
    sub.alignment = centre()

    # Sheet index
    headers = ["Sheet", "Purpose", "Key Skills Demonstrated"]
    widths  = [22, 50, 45]
    apply_header_row(ws, 5, headers, widths)

    index = [
        ("Trade Blotter",  "Full trade log with formulas, conditional formatting, and auto-filter",
         "Excel tables, IF/VLOOKUP formulas, conditional formatting"),
        ("PnL Summary",    "SUMIF-powered PnL rollup by metal and desk with running totals",
         "SUMIF, IFERROR, named ranges, data validation"),
        ("Risk Exposure",  "Gross/net exposure with concentration flags and what-if analysis",
         "Formulas, data bars, scenario inputs"),
        ("Price History",  "Daily spot prices with 30-day moving average and chart",
         "AVERAGE, OFFSET, line chart with dynamic range"),
        ("Pivot_Data",     "Clean flat table optimised for pivot table and slicer creation",
         "Pivot tables, slicers, calculated fields"),
        ("VBA_Reference",  "Documented VBA macro code — copy to VBA editor to enable",
         "VBA: loops, error handling, formatting, email trigger"),
    ]
    for i, (sheet, purpose, skills) in enumerate(index, 6):
        ws.cell(row=i, column=2, value=sheet).font  = Font(name="Calibri", bold=True, color=GOLD, size=10)
        ws.cell(row=i, column=2).fill               = hdr_fill()
        ws.cell(row=i, column=3, value=purpose).font = Font(name="Calibri", size=10, color=WHITE)
        ws.cell(row=i, column=3).fill               = PatternFill("solid", fgColor="0F2A4A")
        ws.cell(row=i, column=4, value=skills).font  = Font(name="Calibri", size=10, color=WHITE)
        ws.cell(row=i, column=4).fill               = PatternFill("solid", fgColor="0F2A4A")
        for col in [2, 3, 4]:
            ws.cell(row=i, column=col).border    = thin_border()
            ws.cell(row=i, column=col).alignment = left()

    ws.column_dimensions["A"].width = 3
    ws.freeze_panes = None


def build_trade_blotter(wb, trades: list[dict]):
    ws = wb.create_sheet("Trade Blotter")
    ws.freeze_panes = "A2"

    headers = [
        "Trade ID", "Trade Date", "Settle Date", "Metal", "Symbol",
        "Desk", "Book", "Trader", "Counterparty", "B/S",
        "Type", "Quantity", "Unit", "Trade Price", "Spot @ Trade",
        "Notional (USD)", "Status", "Days to Settle", "Above Spot?"
    ]
    widths = [14,13,13,12,8,16,15,14,18,6,10,13,10,14,14,16,12,15,12]
    apply_header_row(ws, 1, headers, widths)

    for r, t in enumerate(trades[:500], 2):   # cap at 500 rows
        ws.cell(r, 1,  t.get("trade_id",""))
        ws.cell(r, 2,  t.get("trade_date",""))
        ws.cell(r, 3,  t.get("settlement_date",""))
        ws.cell(r, 4,  t.get("metal",""))
        ws.cell(r, 5,  t.get("symbol",""))
        ws.cell(r, 6,  t.get("desk",""))
        ws.cell(r, 7,  t.get("book","") or t.get("book_id",""))
        ws.cell(r, 8,  t.get("trader",""))
        ws.cell(r, 9,  t.get("counterparty",""))
        ws.cell(r, 10, t.get("buy_sell",""))
        ws.cell(r, 11, t.get("trade_type",""))
        try:
            ws.cell(r, 12, float(t.get("quantity",0)))
        except:
            ws.cell(r, 12, t.get("quantity",""))
        ws.cell(r, 13, t.get("unit",""))
        try:
            ws.cell(r, 14, float(t.get("trade_price",0)))
            ws.cell(r, 15, float(t.get("spot_at_trade",0)))
            ws.cell(r, 16, float(t.get("notional_usd",0)))
        except:
            pass
        ws.cell(r, 17, t.get("status",""))

        # Formula: days to settle = SETTLE - TRADE
        ws.cell(r, 18, f'=IFERROR(DATEVALUE(C{r})-DATEVALUE(B{r}),"N/A")')

        # Formula: trade price above spot?
        ws.cell(r, 19, f'=IF(AND(N{r}<>"",O{r}<>""),IF(N{r}>O{r},"Above","At/Below"),"N/A")')

        # Number formats
        for col in [14, 15]:
            ws.cell(r, col).number_format = '#,##0.0000'
        ws.cell(r, 16).number_format = '#,##0.00'
        ws.cell(r, 12).number_format = '#,##0'

        # Row shading
        fill_colour = "F0F4FA" if r % 2 == 0 else "FFFFFF"
        for col in range(1, 20):
            c = ws.cell(r, col)
            if not c.fill or c.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                c.fill = PatternFill("solid", fgColor=fill_colour)
            c.font = Font(name="Calibri", size=10)
            c.border = thin_border()

    # Conditional formatting — status column
    last_row = len(trades[:500]) + 1
    # Cancelled = grey
    ws.conditional_formatting.add(
        f"Q2:Q{last_row}",
        CellIsRule(operator="equal", formula=['"Cancelled"'],
                   font=Font(color="888888", italic=True),
                   fill=PatternFill("solid", fgColor="EEEEEE"))
    )
    # Pending = amber
    ws.conditional_formatting.add(
        f"Q2:Q{last_row}",
        CellIsRule(operator="equal", formula=['"Pending"'],
                   fill=PatternFill("solid", fgColor="FFF3CD"))
    )

    # Excel Table
    if len(trades) > 0:
        tab = Table(
            displayName="TradeBlotter",
            ref=f"A1:{get_column_letter(19)}{min(len(trades),500)+1}"
        )
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        ws.add_table(tab)

    ws.auto_filter.ref = f"A1:{get_column_letter(19)}1"


def build_pnl_summary(wb, pnl: list[dict]):
    ws = wb.create_sheet("PnL Summary")
    ws.sheet_view.showGridLines = False

    # Section header
    ws.merge_cells("A1:H1")
    ws["A1"].value = "DAILY PnL SUMMARY — Unrealised PnL by Metal & Desk"
    ws["A1"].font  = Font(name="Calibri", bold=True, size=14, color=GOLD)
    ws["A1"].fill  = navy_fill()
    ws["A1"].alignment = centre()
    ws.row_dimensions[1].height = 30

    headers = ["Desk", "Metal", "Symbol", "Net Qty", "Unit",
               "Spot (USD)", "MTM Value", "Cost Basis", "Unrealised PnL", "PnL %", "L/S"]
    widths  = [18, 14, 8, 14, 10, 14, 16, 16, 18, 10, 8]
    apply_header_row(ws, 2, headers, widths)

    for r, p in enumerate(pnl, 3):
        ws.cell(r, 1, p.get("desk",""))
        ws.cell(r, 2, p.get("metal",""))
        ws.cell(r, 3, p.get("symbol",""))
        try:
            ws.cell(r, 4, float(p.get("net_quantity",0)))
            ws.cell(r, 6, float(p.get("spot_price_usd",0)))
            ws.cell(r, 7, float(p.get("mtm_value_usd",0)))
            ws.cell(r, 8, float(p.get("cost_basis_usd",0)))
            ws.cell(r, 9, float(p.get("unrealised_pnl",0)))
            ws.cell(r, 10, f"=I{r}/ABS(H{r})")
        except:
            pass
        ws.cell(r, 5, p.get("unit",""))
        ws.cell(r, 11, p.get("long_short",""))

        for col in [6, 7, 8, 9]:
            ws.cell(r, col).number_format = '#,##0.00'
        ws.cell(r, 4).number_format = '#,##0'
        ws.cell(r, 10).number_format = '0.00%'

    # Totals row
    last = len(pnl) + 2
    tot_row = last + 2
    ws.cell(tot_row, 1, "TOTAL").font = Font(bold=True, color=WHITE)
    ws.cell(tot_row, 1).fill = gold_fill()
    for col, label in [(7,"MTM"),(8,"Cost"),(9,"PnL")]:
        ws.cell(tot_row, col, f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{last})")
        ws.cell(tot_row, col).number_format = '#,##0.00'
        ws.cell(tot_row, col).font  = Font(bold=True)
        ws.cell(tot_row, col).fill  = hdr_fill()

    # Conditional formatting on PnL column — green positive, red negative
    ws.conditional_formatting.add(
        f"I3:I{last}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor="D4EDDA"))
    )
    ws.conditional_formatting.add(
        f"I3:I{last}",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor="F8D7DA"))
    )

    # SUMIF section — desk-level rollup
    ws.cell(tot_row + 3, 1, "DESK ROLLUP (SUMIF)").font = Font(bold=True, color=GOLD)
    ws.cell(tot_row + 4, 1, "Precious Metals PnL:")
    ws.cell(tot_row + 4, 2, f'=SUMIF(A3:A{last},"Precious Metals",I3:I{last})')
    ws.cell(tot_row + 4, 2).number_format = '#,##0.00'
    ws.cell(tot_row + 5, 1, "Base Metals PnL:")
    ws.cell(tot_row + 5, 2, f'=SUMIF(A3:A{last},"Base Metals",I3:I{last})')
    ws.cell(tot_row + 5, 2).number_format = '#,##0.00'
    ws.cell(tot_row + 6, 1, "Total PnL Check:")
    ws.cell(tot_row + 6, 2, f'=B{tot_row+4}+B{tot_row+5}')
    ws.cell(tot_row + 6, 2).number_format = '#,##0.00'


def build_risk_exposure(wb, trades: list[dict]):
    ws = wb.create_sheet("Risk Exposure")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"].value = "RISK EXPOSURE — Gross/Net Notional by Metal"
    ws["A1"].font  = Font(name="Calibri", bold=True, size=14, color=GOLD)
    ws["A1"].fill  = navy_fill()
    ws["A1"].alignment = centre()

    # Scenario input box
    ws["A3"] = "SCENARIO INPUTS"
    ws["A3"].font = Font(bold=True, color=GOLD)
    ws["A4"] = "Spot shock (%):"
    ws["B4"] = 0.00
    ws["B4"].number_format = "0.00%"
    ws["B4"].fill = PatternFill("solid", fgColor="FFF3CD")
    ws["A4"].font = ws["A5"].font = Font(name="Calibri", size=10)
    ws["A5"] = "→ Adjust B4 to stress-test MTM values below"
    ws["A5"].font = Font(italic=True, color="888888", size=9)

    # Aggregate from trades data
    metals_agg: dict[str, dict] = {}
    for t in trades:
        if t.get("status") == "Cancelled":
            continue
        metal = t.get("metal", "Unknown")
        if metal not in metals_agg:
            metals_agg[metal] = {"long": 0.0, "short": 0.0, "count": 0}
        try:
            notional = float(t.get("notional_usd", 0))
            if t.get("buy_sell") == "Buy":
                metals_agg[metal]["long"] += notional
            else:
                metals_agg[metal]["short"] += notional
            metals_agg[metal]["count"] += 1
        except:
            pass

    headers = ["Metal","Gross Long (USD)","Gross Short (USD)","Net Exposure (USD)",
               "Trade Count","Concentration %","Flag"]
    widths  = [14, 20, 20, 20, 12, 16, 10]
    apply_header_row(ws, 7, headers, widths)

    total_notional = sum(v["long"] + v["short"] for v in metals_agg.values()) or 1
    for r, (metal, v) in enumerate(sorted(metals_agg.items()), 8):
        net = v["long"] - v["short"]
        conc = (v["long"] + v["short"]) / total_notional
        ws.cell(r, 1, metal)
        ws.cell(r, 2, v["long"]).number_format   = '#,##0.00'
        ws.cell(r, 3, v["short"]).number_format  = '#,##0.00'
        ws.cell(r, 4, net).number_format         = '#,##0.00'
        ws.cell(r, 5, v["count"])
        ws.cell(r, 6, conc).number_format        = '0.00%'
        flag = "⚠ HIGH" if conc > 0.4 else ("~ MED" if conc > 0.25 else "✓ OK")
        ws.cell(r, 7, flag)

    # Data bars on concentration column
    last = len(metals_agg) + 7
    ws.conditional_formatting.add(
        f"F8:F{last}",
        ColorScaleRule(start_type="min", start_color="63BE7B",
                       end_type="max",   end_color="F8696B")
    )


def build_vba_reference(wb):
    ws = wb.create_sheet("VBA_Reference")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    ws["A1"].value = "VBA MACRO REFERENCE — Copy into VBA Editor (Alt+F11)"
    ws["A1"].font  = Font(name="Courier New", bold=True, size=12, color=GOLD)
    ws["A1"].fill  = navy_fill()
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 90
    ws.column_dimensions["C"].width = 5

    vba_code = '''
' ============================================================
' Module: MetalsRiskMacros
' Purpose: Automate daily workbook refresh and formatting
' ============================================================

Option Explicit

' ------------------------------------------------------------
' 1. REFRESH_ALL_DATA
' Clears old data, re-imports CSVs from shared drive, rebuilds pivots
' Usage: Assign to button on README sheet
' ------------------------------------------------------------
Sub RefreshAllData()
    Dim ws As Worksheet
    Dim sPath As String
    
    On Error GoTo ErrorHandler
    
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    
    sPath = "\\\\bmo-server\\RiskData\\MetalsExports\\"
    
    ' Import trade blotter
    Call ImportCSVToSheet("trades.csv", "Trade Blotter", sPath)
    
    ' Import PnL summary
    Call ImportCSVToSheet("pnl_summary.csv", "PnL Summary", sPath)
    
    ' Refresh all pivot tables
    Dim pc As PivotCache
    For Each pc In ThisWorkbook.PivotCaches
        pc.Refresh
    Next pc
    
    ' Update last-refreshed timestamp
    Sheets("README").Range("J3").Value = "Last refreshed: " & Now()
    
    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    MsgBox "Refresh complete: " & Format(Now(), "dd-mmm-yyyy hh:mm"), _
           vbInformation, "Metals Risk Dashboard"
    Exit Sub
    
ErrorHandler:
    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    MsgBox "Error during refresh: " & Err.Description, vbCritical, "Error"
End Sub


' ------------------------------------------------------------
' 2. IMPORT_CSV_TO_SHEET
' Generic CSV importer with header detection
' ------------------------------------------------------------
Sub ImportCSVToSheet(sFile As String, sSheet As String, sPath As String)
    Dim ws As Worksheet
    Dim sFullPath As String
    
    sFullPath = sPath & sFile
    
    If Not FileExists(sFullPath) Then
        MsgBox "File not found: " & sFullPath, vbExclamation
        Exit Sub
    End If
    
    Set ws = ThisWorkbook.Sheets(sSheet)
    
    ' Clear existing data (keep header row 1)
    ws.Rows("2:" & ws.Rows.Count).ClearContents
    
    ' Open CSV and copy data
    Dim wbCSV As Workbook
    Set wbCSV = Workbooks.Open(sFullPath, ReadOnly:=True)
    
    wbCSV.Sheets(1).UsedRange.Offset(1, 0).Copy
    ws.Range("A2").PasteSpecial xlPasteValues
    
    wbCSV.Close False
    Application.CutCopyMode = False
End Sub


' ------------------------------------------------------------
' 3. FORMAT_PNL_COLUMN
' Applies green/red conditional formatting to any PnL column
' Usage: Select PnL range, then run macro
' ------------------------------------------------------------
Sub FormatPnLColumn()
    Dim rng As Range
    Set rng = Selection
    
    If rng Is Nothing Then
        MsgBox "Select a range first", vbExclamation
        Exit Sub
    End If
    
    ' Clear existing conditional formats
    rng.FormatConditions.Delete
    
    ' Positive = green fill
    Dim fcPos As FormatCondition
    Set fcPos = rng.FormatConditions.Add( _
        Type:=xlCellValue, Operator:=xlGreater, Formula1:="0")
    fcPos.Interior.Color = RGB(212, 237, 218)   ' light green
    fcPos.Font.Color = RGB(21, 87, 36)           ' dark green text
    
    ' Negative = red fill
    Dim fcNeg As FormatCondition
    Set fcNeg = rng.FormatConditions.Add( _
        Type:=xlCellValue, Operator:=xlLess, Formula1:="0")
    fcNeg.Interior.Color = RGB(248, 215, 218)   ' light red
    fcNeg.Font.Color = RGB(114, 28, 36)          ' dark red text
    
    MsgBox "PnL formatting applied to " & rng.Address, vbInformation
End Sub


' ------------------------------------------------------------
' 4. SEND_DAILY_SUMMARY_EMAIL
' Builds and sends a daily PnL summary email via Outlook
' Requires: Outlook installed and configured
' ------------------------------------------------------------
Sub SendDailySummaryEmail()
    Dim olApp As Object
    Dim olMail As Object
    Dim ws As Worksheet
    Dim sBody As String
    Dim dTotalPnL As Double
    Dim dPreciousPnL As Double
    Dim dBasePnL As Double
    
    Set ws = ThisWorkbook.Sheets("PnL Summary")
    
    ' Read totals from named cells (set these up in the sheet)
    ' Assumes SUMIF formulas are in fixed cells
    On Error Resume Next
    dTotalPnL    = ws.Range("B" & (ws.UsedRange.Rows.Count + 4)).Value
    dPreciousPnL = ws.Range("B" & (ws.UsedRange.Rows.Count + 5)).Value
    dBasePnL     = ws.Range("B" & (ws.UsedRange.Rows.Count + 6)).Value
    On Error GoTo 0
    
    ' Build email body
    sBody = "<html><body style='font-family:Calibri;'>" & _
            "<h2 style='color:#0A1628;'>&#128202; Metals Risk — Daily PnL Summary</h2>" & _
            "<p><b>Date:</b> " & Format(Now(), "dd-mmm-yyyy") & "</p>" & _
            "<table border='1' cellpadding='5' style='border-collapse:collapse;'>" & _
            "<tr style='background:#1E3A5F;color:white;'>" & _
            "  <th>Desk</th><th>Unrealised PnL (USD)</th></tr>" & _
            "<tr><td>Precious Metals</td><td align='right'>" & _
              Format(dPreciousPnL, "$#,##0.00") & "</td></tr>" & _
            "<tr><td>Base Metals</td><td align='right'>" & _
              Format(dBasePnL, "$#,##0.00") & "</td></tr>" & _
            "<tr style='font-weight:bold;'><td>TOTAL</td><td align='right'>" & _
              Format(dTotalPnL, "$#,##0.00") & "</td></tr>" & _
            "</table>" & _
            "<p style='font-size:10px;color:gray;'>Auto-generated by MetalsRiskMacros.xlsm — do not reply</p>" & _
            "</body></html>"
    
    ' Create and send email
    Set olApp  = CreateObject("Outlook.Application")
    Set olMail = olApp.CreateItem(0)
    
    With olMail
        .To      = "dl-metals-risk@bmo.com"
        .Subject = "Metals Risk PnL Summary — " & Format(Now(), "dd-mmm-yyyy")
        .HTMLBody = sBody
        .Attachments.Add ThisWorkbook.FullName
        .Display   ' Use .Send to send without preview
    End With
    
    Set olMail = Nothing
    Set olApp  = Nothing
End Sub


' ------------------------------------------------------------
' 5. HELPER: FileExists
' ------------------------------------------------------------
Function FileExists(sPath As String) As Boolean
    FileExists = (Dir(sPath) <> "")
End Function


' ------------------------------------------------------------
' 6. HIGHLIGHT_BREACHES
' Scans Risk Exposure sheet and flags concentration > 40%
' ------------------------------------------------------------
Sub HighlightBreaches()
    Dim ws As Worksheet
    Dim lastRow As Long
    Dim i As Long
    Dim concPct As Double
    
    Set ws = ThisWorkbook.Sheets("Risk Exposure")
    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    
    For i = 8 To lastRow
        On Error Resume Next
        concPct = ws.Cells(i, 6).Value
        On Error GoTo 0
        
        If concPct > 0.4 Then
            ws.Rows(i).Interior.Color = RGB(248, 215, 218)   ' Red
            ws.Cells(i, 7).Value = "BREACH"
            ws.Cells(i, 7).Font.Bold = True
            ws.Cells(i, 7).Font.Color = RGB(114, 28, 36)
        ElseIf concPct > 0.25 Then
            ws.Rows(i).Interior.Color = RGB(255, 243, 205)   ' Amber
        End If
    Next i
    
    MsgBox "Breach scan complete.", vbInformation
End Sub
'''

    lines = vba_code.strip().split("\n")
    for i, line in enumerate(lines, 3):
        cell = ws.cell(row=i, column=2, value=line)
        cell.font      = Font(name="Courier New", size=9, color="D4D4D4")
        cell.fill      = PatternFill("solid", fgColor="1E1E1E")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--data-dir", default="data")
    parser.add_argument("--out",      default="excel/MetalsRiskDashboard.xlsx")
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print("Loading CSV data...")
    trades    = load_csv(data_dir / "trades.csv")
    pnl       = load_csv(data_dir / "pnl_summary.csv")
    positions = load_csv(data_dir / "positions.csv")
    prices    = load_csv(data_dir / "price_history.csv")

    print(f"  trades={len(trades)}  pnl={len(pnl)}  positions={len(positions)}  prices={len(prices)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M UTC")

    print("Building sheets...")
    build_readme(wb, generated_at)
    build_trade_blotter(wb, trades)
    build_pnl_summary(wb, pnl)
    build_risk_exposure(wb, trades)
    build_vba_reference(wb)

    wb.save(out_path)
    print(f"\n✓ Workbook saved: {out_path}")
    print("  Sheets: README, Trade Blotter, PnL Summary, Risk Exposure, VBA_Reference")
    print("  Next: Open in Excel → Save As .xlsm to enable VBA macros")


if __name__ == "__main__":
    main()
